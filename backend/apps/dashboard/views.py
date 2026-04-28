from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models import Avg, Count, ExpressionWrapper, F, Q, Sum, DurationField
from django.db.models.functions import TruncWeek
from django.utils import timezone
from datetime import timedelta
from apps.jobs.models import Job, JobCollaborator
from apps.candidates.models import CandidateJobMapping, Referral, MACRO_STAGE_CHOICES
from apps.requisitions.models import Requisition
from apps.interviews.models import Interview
from apps.core.permissions import rbac_perm
from apps.core.rbac import has_permission


def _jobs_scope(user, jobs_qs):
    """Scope a Job queryset to what the user is allowed to see in the dashboard."""
    if has_permission(user, 'MANAGE_USERS'):
        return jobs_qs
    if has_permission(user, 'SCHEDULE_INTERVIEW'):
        collab_job_ids = JobCollaborator.objects.filter(user=user).values_list('job_id', flat=True)
        return jobs_qs.filter(Q(created_by=user) | Q(id__in=collab_job_ids))
    if has_permission(user, 'APPROVE_REQUISITIONS'):
        return jobs_qs.filter(hiring_manager=user)
    # Interviewers and custom roles with VIEW_REPORTS but no job-management perms see all
    return jobs_qs


class DashboardExcelReportView(APIView):
    permission_classes = [rbac_perm('VIEW_REPORTS')]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from io import BytesIO

        user = request.user

        jobs_qs = _jobs_scope(user, Job.objects.all())

        # Accept the same filter params
        for param, field in [
            ('departments', 'department_id__in'),
            ('hiring_managers', 'hiring_manager_id__in'),
            ('locations', 'location__in'),
            ('designations', 'title__in'),
        ]:
            val = request.query_params.get(param, '')
            if val:
                jobs_qs = jobs_qs.filter(**{field: [v for v in val.split(',') if v]})

        job_ids = list(jobs_qs.values_list('id', flat=True))
        mappings = CandidateJobMapping.objects.filter(job_id__in=job_ids)

        stage_counts = dict(
            mappings.values('macro_stage').annotate(c=Count('id')).values_list('macro_stage', 'c')
        )
        total_applies = sum(stage_counts.values())
        total_views = jobs_qs.aggregate(total=Sum('view_count'))['total'] or 0
        interviews_qs = Interview.objects.filter(mapping__job_id__in=job_ids)

        now = timezone.now()
        avg_duration = (
            mappings.filter(macro_stage__in=['OFFERED', 'JOINED'])
            .annotate(duration=ExpressionWrapper(F('stage_updated_at') - F('created_at'), output_field=DurationField()))
            .aggregate(avg=Avg('duration'))['avg']
        )
        avg_days_to_hire = round(avg_duration.days) if avg_duration else None

        def source_agg(qs):
            return qs.aggregate(
                referral=Count('id', filter=Q(candidate__source__in=['referral'])),
                recruiter_sourced=Count('id', filter=Q(candidate__source__in=['recruiter_upload'])),
                inbound=Count('id', filter=Q(candidate__source__in=['linkedin', 'naukri'])),
                admin=Count('id', filter=Q(candidate__source__in=['manual'])),
            )

        # ── Styles ────────────────────────────────────────────────────────────
        HEADER_FILL  = PatternFill('solid', fgColor='0058BE')
        SECTION_FILL = PatternFill('solid', fgColor='DAE2FA')
        HEADER_FONT  = Font(bold=True, color='FFFFFF', size=11)
        SECTION_FONT = Font(bold=True, color='0058BE', size=10)
        CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        thin   = Side(style='thin', color='C2C6D6')
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(ws, row, cols):
            for c in range(1, cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER
                cell.border = BORDER

        def write_row(ws, row, values):
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.alignment = LEFT
                cell.border = BORDER

        def kv_row(ws, row, label, value):
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = Font(bold=True, size=10)
            lc.fill = SECTION_FILL
            lc.border = BORDER
            vc = ws.cell(row=row, column=2, value=str(value) if value is not None else '—')
            vc.alignment = LEFT
            vc.border = BORDER

        wb = openpyxl.Workbook()

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 1 — Overview
        # ══════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Overview'
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20

        ws1.merge_cells('A1:B1')
        ws1['A1'] = 'DASHBOARD OVERVIEW'
        ws1['A1'].fill = HEADER_FILL
        ws1['A1'].font = HEADER_FONT
        ws1['A1'].alignment = CENTER
        ws1.row_dimensions[1].height = 24

        total_offered = stage_counts.get('OFFERED', 0)
        total_joined  = stage_counts.get('JOINED', 0)
        ever_offered  = total_offered + total_joined
        offer_join_rate = round(total_joined / ever_offered * 100) if ever_offered else 0

        overview_rows = [
            ('Report Generated',  now.strftime('%d-%b-%Y %H:%M UTC')),
            ('Active Jobs',       jobs_qs.count()),
            ('Total Views',       total_views),
            ('Total Applications',total_applies),
            ('Applied (pending)', stage_counts.get('APPLIED', 0)),
            ('Shortlisted',       stage_counts.get('SHORTLISTED', 0)),
            ('Interview Stage',   stage_counts.get('INTERVIEW', 0)),
            ('Offered',           total_offered),
            ('Joined',            total_joined),
            ('Dropped',           stage_counts.get('DROPPED', 0)),
            ('Total Interviews',  interviews_qs.count()),
            ('Avg Days to Hire',  avg_days_to_hire if avg_days_to_hire is not None else '—'),
            ('Offer-to-Join Rate',f'{offer_join_rate}%'),
        ]
        for r, (label, val) in enumerate(overview_rows, 2):
            kv_row(ws1, r, label, val)

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 2 — Recruitment Funnel
        # ══════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Recruitment Funnel')
        for col, w in zip(['A', 'B', 'C'], [20, 12, 20]):
            ws2.column_dimensions[col].width = w

        funnel_headers = ['Stage', 'Candidates', 'Conversion from Applied (%)']
        style_header(ws2, 1, len(funnel_headers))
        write_row(ws2, 1, funnel_headers)

        app_total = (
            stage_counts.get('APPLIED', 0) + stage_counts.get('SHORTLISTED', 0) +
            stage_counts.get('INTERVIEW', 0) + total_offered + total_joined +
            stage_counts.get('DROPPED', 0)
        )
        funnel_data = [
            ('Applied',     app_total),
            ('Shortlisted', stage_counts.get('SHORTLISTED', 0) + stage_counts.get('INTERVIEW', 0) + total_offered + total_joined),
            ('Interview',   stage_counts.get('INTERVIEW', 0) + total_offered + total_joined),
            ('Offered',     total_offered + total_joined),
            ('Joined',      total_joined),
        ]
        for r, (stage, count) in enumerate(funnel_data, 2):
            conv = round(count / app_total * 100, 1) if app_total else 0
            write_row(ws2, r, [stage, count, f'{conv}%'])

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 3 — Source Breakdown
        # ══════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('Source Breakdown')
        for col, w in zip(['A', 'B', 'C', 'D', 'E'], [20, 16, 16, 16, 16]):
            ws3.column_dimensions[col].width = w

        src_headers = ['Source', 'All Candidates', 'Progressed', 'Offered', 'Joined']
        style_header(ws3, 1, len(src_headers))
        write_row(ws3, 1, src_headers)

        all_src  = source_agg(mappings)
        prog_src = source_agg(mappings.exclude(macro_stage__in=['APPLIED', 'DROPPED']))
        off_src  = source_agg(mappings.filter(macro_stage='OFFERED'))
        join_src = source_agg(mappings.filter(macro_stage='JOINED'))

        src_map = [
            ('Referral',         'referral'),
            ('Recruiter Sourced','recruiter_sourced'),
            ('Inbound',          'inbound'),
            ('Admin',            'admin'),
        ]
        for r, (label, key) in enumerate(src_map, 2):
            write_row(ws3, r, [label, all_src[key], prog_src[key], off_src[key], join_src[key]])

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 4 — Pipeline by Job
        # ══════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet('Pipeline by Job')
        job_headers = [
            'Job Code', 'Title', 'Department', 'Status', 'Location',
            'Applied', 'Shortlisted', 'Interview', 'Offered', 'Joined', 'Dropped', 'Total',
        ]
        col_widths4 = [14, 28, 20, 12, 16, 10, 12, 12, 10, 10, 10, 10]
        for i, w in enumerate(col_widths4, 1):
            ws4.column_dimensions[ws4.cell(1, i).column_letter].width = w

        style_header(ws4, 1, len(job_headers))
        write_row(ws4, 1, job_headers)

        jobs_detail = (
            jobs_qs
            .select_related('department')
            .prefetch_related('candidate_mappings')
            .order_by('job_code')
        )
        for r, job in enumerate(jobs_detail, 2):
            jm = CandidateJobMapping.objects.filter(job=job)
            jsc = dict(jm.values('macro_stage').annotate(c=Count('id')).values_list('macro_stage', 'c'))
            total_j = sum(jsc.values())
            write_row(ws4, r, [
                job.job_code or '—',
                job.title,
                job.department.name if job.department else '—',
                job.status,
                job.location or '—',
                jsc.get('APPLIED', 0),
                jsc.get('SHORTLISTED', 0),
                jsc.get('INTERVIEW', 0),
                jsc.get('OFFERED', 0),
                jsc.get('JOINED', 0),
                jsc.get('DROPPED', 0),
                total_j,
            ])

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 5 — Pending Actions
        # ══════════════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet('Pending Actions')
        ws5.column_dimensions['A'].width = 30
        ws5.column_dimensions['B'].width = 16

        ws5.merge_cells('A1:B1')
        ws5['A1'] = 'PENDING ACTIONS'
        ws5['A1'].fill = HEADER_FILL
        ws5['A1'].font = HEADER_FONT
        ws5['A1'].alignment = CENTER
        ws5.row_dimensions[1].height = 24

        pending_approvals = Requisition.objects.filter(status='pending_approval').count() if has_permission(user, 'APPROVE_REQUISITIONS') else 0
        pending_feedback  = Interview.objects.filter(
            interviewer=user, status='scheduled', scheduled_at__lt=now
        ).count()
        pending_referrals = Referral.objects.filter(status='pending').count() if has_permission(user, 'MANAGE_USERS') else 0
        stale_threshold   = now - timedelta(days=7)
        stale_candidates  = CandidateJobMapping.objects.filter(
            job_id__in=job_ids,
            stage_updated_at__lt=stale_threshold,
            macro_stage__in=['APPLIED', 'SHORTLISTED', 'INTERVIEW'],
        ).count()

        for r, (label, val) in enumerate([
            ('Pending Approvals',  pending_approvals),
            ('Pending Feedback',   pending_feedback),
            ('Pending Referrals',  pending_referrals),
            ('Stale Candidates (7d)', stale_candidates),
        ], 2):
            kv_row(ws5, r, label, val)

        # ── Serialize and return ──────────────────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Dashboard_Report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


def _weekly_history(queryset, date_field='created_at'):
    """Return list of {name, value} dicts bucketed by ISO week start."""
    rows = (
        queryset
        .annotate(week=TruncWeek(date_field))
        .values('week')
        .annotate(c=Count('id'))
        .order_by('week')
    )
    return [{'name': f"{row['week'].day}-{row['week'].strftime('%b-%Y')}", 'value': row['c']} for row in rows]


def _calc_trend(queryset, now, date_field='created_at'):
    """Return (pct, trend_up) comparing this 7-day window vs previous 7-day window."""
    week_start = now - timedelta(days=7)
    prev_start = now - timedelta(days=14)
    this_week = queryset.filter(**{f'{date_field}__gte': week_start}).count()
    last_week = queryset.filter(**{f'{date_field}__gte': prev_start, f'{date_field}__lt': week_start}).count()
    if last_week == 0:
        pct = 100 if this_week > 0 else 0
        trend_up = True if this_week > 0 else None
    else:
        pct = round(abs((this_week - last_week) / last_week * 100))
        trend_up = this_week >= last_week
    return pct, trend_up


class DashboardSummaryView(APIView):
    permission_classes = [rbac_perm('VIEW_REPORTS')]

    def get(self, request):
        user = request.user
        department = request.query_params.get('department')

        jobs_qs = _jobs_scope(user, Job.objects.all())

        if department:
            jobs_qs = jobs_qs.filter(department_id=department)

        departments_param    = request.query_params.get('departments', '')
        hiring_managers_param = request.query_params.get('hiring_managers', '')
        locations_param      = request.query_params.get('locations', '')
        designations_param   = request.query_params.get('designations', '')

        if departments_param:
            jobs_qs = jobs_qs.filter(department_id__in=[v for v in departments_param.split(',') if v])
        if hiring_managers_param:
            jobs_qs = jobs_qs.filter(hiring_manager_id__in=[v for v in hiring_managers_param.split(',') if v])
        if locations_param:
            jobs_qs = jobs_qs.filter(location__in=[v for v in locations_param.split(',') if v])
        if designations_param:
            jobs_qs = jobs_qs.filter(title__in=[v for v in designations_param.split(',') if v])

        job_ids = list(jobs_qs.values_list('id', flat=True))
        mappings = CandidateJobMapping.objects.filter(job_id__in=job_ids)

        # One GROUP BY query instead of one COUNT per stage (was 8+ queries).
        stage_counts = dict(
            mappings.values('macro_stage').annotate(c=Count('id')).values_list('macro_stage', 'c')
        )
        total_applies = sum(stage_counts.values())

        # SQL SUM instead of fetching all rows and summing in Python.
        total_views = jobs_qs.aggregate(total=Sum('view_count'))['total'] or 0

        now = timezone.now()

        interviews_qs = Interview.objects.filter(mapping__job_id__in=job_ids)
        interviews_count = interviews_qs.count()

        # Weekly history + trend per metric
        applies_history = _weekly_history(mappings)
        applies_trend = _calc_trend(mappings, now)

        jobs_history = _weekly_history(jobs_qs)
        jobs_trend = _calc_trend(jobs_qs, now)

        interviews_history = _weekly_history(interviews_qs, date_field='scheduled_at')
        interviews_trend = _calc_trend(interviews_qs, now, date_field='scheduled_at')

        stage_history = {}
        stage_trend = {}
        for key, _ in MACRO_STAGE_CHOICES:
            qs = mappings.filter(macro_stage=key)
            stage_history[key] = _weekly_history(qs)
            stage_trend[key] = _calc_trend(qs, now)

        def make_metric(title, value, history, trend_tuple):
            pct, up = trend_tuple
            return {'title': title, 'value': value, 'history': history, 'trend_pct': pct, 'trend_up': up}

        metrics = [
            make_metric('Jobs',        jobs_qs.count(),                         jobs_history,                      jobs_trend),
            make_metric('Views',       total_views,                             applies_history,                   applies_trend),
            make_metric('Applies',     total_applies,                           applies_history,                   applies_trend),
            make_metric('Applied',     stage_counts.get('APPLIED', 0),         stage_history['APPLIED'],          stage_trend['APPLIED']),
            make_metric('Shortlisted', stage_counts.get('SHORTLISTED', 0),     stage_history['SHORTLISTED'],      stage_trend['SHORTLISTED']),
            make_metric('Interviews',  interviews_count,                        interviews_history,                interviews_trend),
            make_metric('Interview',   stage_counts.get('INTERVIEW', 0),       stage_history['INTERVIEW'],        stage_trend['INTERVIEW']),
            make_metric('Offered',     stage_counts.get('OFFERED', 0),         stage_history['OFFERED'],          stage_trend['OFFERED']),
            make_metric('Joined',      stage_counts.get('JOINED', 0),          stage_history['JOINED'],           stage_trend['JOINED']),
            make_metric('Dropped',     stage_counts.get('DROPPED', 0),         stage_history['DROPPED'],          stage_trend['DROPPED']),
        ]

        all_candidates = total_applies
        progressed = total_applies - stage_counts.get('APPLIED', 0) - stage_counts.get('DROPPED', 0)

        # Average days from application to offer/join
        avg_duration = (
            mappings.filter(macro_stage__in=['OFFERED', 'JOINED'])
            .annotate(duration=ExpressionWrapper(F('stage_updated_at') - F('created_at'), output_field=DurationField()))
            .aggregate(avg=Avg('duration'))['avg']
        )
        avg_days_to_hire = round(avg_duration.days) if avg_duration else None

        def source_breakdown(qs):
            # One aggregate query instead of 4 separate COUNT queries per call.
            agg = qs.aggregate(
                referral=Count('id', filter=Q(candidate__source__in=['referral'])),
                recruiter_sourced=Count('id', filter=Q(candidate__source__in=['recruiter_upload'])),
                inbound=Count('id', filter=Q(candidate__source__in=['linkedin', 'naukri'])),
                partner=Count('id', filter=Q(candidate__source__in=['manual'])),
            )
            return {
                'Referral': agg['referral'],
                'Recruiter Sourced': agg['recruiter_sourced'],
                'Inbound': agg['inbound'],
                'Admin': agg['partner'],
            }

        return Response({
            'metrics': metrics,
            'avg_days_to_hire': avg_days_to_hire,
            'recruitment_progress': {
                'all_candidates': all_candidates,
                'progressed_candidates': progressed,
                'all_breakdown': source_breakdown(mappings),
                'progressed_breakdown': source_breakdown(mappings.exclude(macro_stage__in=['APPLIED', 'DROPPED'])),
                'offered_breakdown': source_breakdown(mappings.filter(macro_stage='OFFERED')),
                'joined_breakdown': source_breakdown(mappings.filter(macro_stage='JOINED')),
                'pipeline_stages': [
                    {'label': 'Shortlisted', 'value': stage_counts.get('SHORTLISTED', 0)},
                    {'label': 'Interview',   'value': stage_counts.get('INTERVIEW', 0)},
                    {'label': 'Applied',     'value': stage_counts.get('APPLIED', 0)},
                ],
            },
        })


class DashboardFilterOptionsView(APIView):
    permission_classes = [rbac_perm('VIEW_REPORTS')]

    def get(self, request):
        from apps.departments.models import Department
        from apps.accounts.models import User

        departments = list(Department.objects.values('id', 'name').order_by('name'))
        hiring_managers = list(
            User.objects.filter(is_active=True)
            .values('id', 'full_name', 'role')
            .order_by('full_name')
        )
        locations = list(
            Job.objects.exclude(location='')
            .values_list('location', flat=True)
            .distinct()
            .order_by('location')
        )
        designations = list(
            Job.objects.values_list('title', flat=True)
            .distinct()
            .order_by('title')
        )
        return Response({
            'departments': departments,
            'hiring_managers': hiring_managers,
            'locations': locations,
            'designations': designations,
        })


class DashboardFunnelView(APIView):
    permission_classes = [rbac_perm('VIEW_REPORTS')]

    def get(self, request):
        department = request.query_params.get('department')
        jobs_qs = Job.objects.filter(status='open')
        if department:
            jobs_qs = jobs_qs.filter(department_id=department)
        job_ids = jobs_qs.values_list('id', flat=True)
        mappings = CandidateJobMapping.objects.filter(job_id__in=job_ids)

        # One GROUP BY query instead of one COUNT per stage.
        stage_counts = dict(
            mappings.values('macro_stage').annotate(c=Count('id')).values_list('macro_stage', 'c')
        )
        funnel = [
            {'stage': key, 'label': label, 'count': stage_counts.get(key, 0)}
            for key, label in MACRO_STAGE_CHOICES
        ]
        return Response(funnel)


class DashboardPendingActionsView(APIView):
    permission_classes = [rbac_perm('VIEW_REPORTS')]

    def get(self, request):
        user = request.user

        if has_permission(user, 'APPROVE_REQUISITIONS'):
            pending_approvals = Requisition.objects.filter(status='pending_approval').count()
        elif has_permission(user, 'MANAGE_REQUISITIONS'):
            pending_approvals = Requisition.objects.filter(status='pending_approval', created_by=user).count()
        else:
            pending_approvals = 0

        now = timezone.now()
        pending_feedback = Interview.objects.filter(
            interviewer=user, status='scheduled', scheduled_at__lt=now
        ).count()

        pending_referrals = (
            Referral.objects.filter(status='pending').count()
            if has_permission(user, 'MANAGE_USERS') else 0
        )

        jobs_qs = _jobs_scope(user, Job.objects.all())

        stale_threshold = now - timedelta(days=7)
        stale_candidates = CandidateJobMapping.objects.filter(
            job_id__in=jobs_qs.values_list('id', flat=True),
            stage_updated_at__lt=stale_threshold,
            macro_stage__in=['APPLIED', 'SHORTLISTED', 'INTERVIEW'],
        ).count()

        return Response({
            'pending_approvals': pending_approvals,
            'pending_feedback': pending_feedback,
            'pending_referrals': pending_referrals,
            'stale_candidates': stale_candidates,
        })
