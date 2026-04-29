from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models import Count, Q
from .models import Job, JobCollaborator, JobHistory
from .serializers import (
    JobListSerializer, JobDetailSerializer,
    JobCollaboratorSerializer, JobHistorySerializer,
)
from apps.candidates.models import CandidateJobMapping
from apps.candidates.serializers import CandidateJobMappingSerializer
from apps.core.permissions import CanEditJob, rbac_perm


def log_job_history(job, event_type, changed_by, description, previous_value=None, new_value=None):
    JobHistory.objects.create(
        job=job,
        event_type=event_type,
        changed_by=changed_by,
        description=description,
        previous_value=previous_value,
        new_value=new_value,
    )


class JobListView(generics.ListAPIView):
    serializer_class = JobListSerializer
    permission_classes = [rbac_perm('VIEW_JOBS')]
    search_fields = ['title', 'job_code', 'location']
    filterset_fields = ['status', 'department', 'hiring_manager', 'location']
    ordering_fields = ['created_at', 'title']

    def get_queryset(self):
        # All four counts are computed as conditional COUNTs in one SQL query — no per-row queries.
        qs = Job.objects.select_related('department', 'hiring_manager', 'requisition').annotate(
            applies_count=Count('candidate_mappings', distinct=True),
            shortlists_count=Count(
                'candidate_mappings',
                filter=Q(candidate_mappings__macro_stage__in=[
                    'SHORTLISTED', 'INTERVIEW', 'OFFERED', 'JOINED'
                ]),
                distinct=True,
            ),
            interviews_count=Count(
                'candidate_mappings',
                filter=Q(candidate_mappings__macro_stage__in=[
                    'INTERVIEW', 'OFFERED', 'JOINED'
                ]),
                distinct=True,
            ),
            offers_count=Count(
                'candidate_mappings',
                filter=Q(candidate_mappings__macro_stage__in=['OFFERED', 'JOINED']),
                distinct=True,
            ),
            joined_count=Count(
                'candidate_mappings',
                filter=Q(candidate_mappings__macro_stage='JOINED'),
                distinct=True,
            ),
        )
        tab = self.request.query_params.get('tab', 'all')
        user = self.request.user
        if tab == 'mine':
            qs = qs.filter(
                Q(created_by=user) | Q(collaborators__user=user) | Q(hiring_manager=user)
            ).distinct()

        date_from = self.request.query_params.get('date_from')
        date_to   = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        return qs


class JobDetailView(generics.RetrieveUpdateAPIView):
    queryset = Job.objects.select_related(
        'department', 'hiring_manager', 'created_by', 'requisition', 'requisition__created_by'
    ).prefetch_related('collaborators__user', 'history__changed_by')
    serializer_class = JobDetailSerializer
    permission_classes = [rbac_perm('VIEW_JOBS'), CanEditJob]

    def perform_update(self, serializer):
        instance = serializer.instance
        old_status = instance.status
        old_title = instance.title
        old_dept_id = str(instance.department_id)
        old_hm_id = str(instance.hiring_manager_id)
        old_jd = instance.job_description

        updated = serializer.save()
        user = self.request.user

        if old_status != updated.status:
            log_job_history(
                updated, 'status_changed', user,
                f'Status changed: {old_status.capitalize()} → {updated.status.capitalize()}',
                previous_value={'status': old_status},
                new_value={'status': updated.status},
            )

        if old_title != updated.title:
            log_job_history(
                updated, 'title_updated', user,
                f'Title updated: "{old_title}" → "{updated.title}"',
                previous_value={'title': old_title},
                new_value={'title': updated.title},
            )

        if old_dept_id != str(updated.department_id):
            log_job_history(
                updated, 'department_changed', user,
                'Department changed',
                previous_value={'department': old_dept_id},
                new_value={'department': str(updated.department_id)},
            )

        if old_hm_id != str(updated.hiring_manager_id):
            log_job_history(
                updated, 'hiring_manager_changed', user,
                'Hiring Manager changed',
                previous_value={'hiring_manager': old_hm_id},
                new_value={'hiring_manager': str(updated.hiring_manager_id)},
            )

        if old_jd != updated.job_description:
            log_job_history(updated, 'jd_updated', user, 'Job description updated')


class JobDeleteView(APIView):
    permission_classes = [rbac_perm('EDIT_JOBS')]

    def delete(self, request, pk):
        job = generics.get_object_or_404(Job, pk=pk)
        job.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class JobPipelineView(APIView):
    """
    GET /jobs/{pk}/pipeline/?stage=APPLIED&include_progressed=true

    Returns active candidates for the given stage first (sorted by priority DESC,
    stage_updated_at DESC), followed by dimmed candidates who have progressed past
    this stage (sorted by stage_updated_at DESC).

    Each record includes:
      - All CandidateJobMapping fields
      - is_current_stage: bool
      - can_move_next: bool
      - latest_round: {round_name, round_status, round_result} (if in INTERVIEW)
    """
    permission_classes = [rbac_perm('VIEW_JOBS')]

    def get(self, request, pk):
        from apps.candidates.models import STAGE_ORDER, PipelineStageHistory
        from apps.candidates.serializers import CandidateJobMappingSerializer

        job = generics.get_object_or_404(Job, pk=pk)
        stage_param = request.query_params.get('stage', '').strip()
        include_progressed = request.query_params.get('include_progressed', 'false').lower() == 'true'

        base_qs = CandidateJobMapping.objects.filter(job=job, is_archived=False).select_related(
            'candidate', 'job'
        ).prefetch_related('interviews')

        # Priority sort order helper
        PRIORITY_ORDER = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}

        def sort_key_active(m):
            return (PRIORITY_ORDER.get(m.priority, 1), -m.stage_updated_at.timestamp())

        if stage_param:
            active_qs = list(base_qs.filter(macro_stage=stage_param))
            active_qs.sort(key=sort_key_active)
        else:
            active_qs = list(base_qs.order_by('-stage_updated_at'))

        dimmed = []
        if include_progressed and stage_param:
            requested_order = STAGE_ORDER.get(stage_param, -1)
            # Candidates who passed through this stage but are now at a higher stage
            past_mapping_ids = set(
                PipelineStageHistory.objects.filter(
                    mapping__job=job,
                    from_macro_stage=stage_param,
                ).values_list('mapping_id', flat=True)
            )
            active_ids = {m.id for m in active_qs}
            dimmed = list(
                base_qs.filter(id__in=past_mapping_ids).exclude(
                    macro_stage=stage_param
                ).exclude(id__in=active_ids)
            )
            dimmed.sort(key=lambda m: -m.stage_updated_at.timestamp())

        # Batch-fetch active reminders for current user — one query, no N+1
        from apps.candidates.models import CandidateReminder
        all_mappings = active_qs + dimmed
        candidate_ids = {m.candidate_id for m in all_mappings}
        reminded_ids = set(
            CandidateReminder.objects.filter(
                candidate_id__in=candidate_ids,
                created_by=request.user,
                is_done=False,
            ).values_list('candidate_id', flat=True)
        )

        results = []
        for mapping in active_qs:
            data = CandidateJobMappingSerializer(mapping).data
            data['is_current_stage'] = True
            data['can_move_next'] = mapping.macro_stage not in ('JOINED', 'DROPPED')
            data['latest_round'] = _get_latest_round(mapping)
            data['has_active_reminder'] = mapping.candidate_id in reminded_ids
            results.append(data)

        for mapping in dimmed:
            data = CandidateJobMappingSerializer(mapping).data
            data['is_current_stage'] = False
            data['can_move_next'] = False
            data['latest_round'] = _get_latest_round(mapping)
            data['has_active_reminder'] = mapping.candidate_id in reminded_ids
            results.append(data)

        return Response(results)


def _get_latest_round(mapping):
    interview = mapping.interviews.order_by('-created_at').first()
    if not interview:
        return None
    return {
        'id': str(interview.id),
        'round_name': interview.round_name,
        'round_status': interview.round_status,
        'round_result': interview.round_result,
        'scheduled_at': interview.scheduled_at.isoformat() if interview.scheduled_at else None,
    }


class JobPipelineStatsView(APIView):
    permission_classes = [rbac_perm('VIEW_JOBS')]

    def get(self, request, pk):
        from apps.candidates.models import MACRO_STAGE_CHOICES
        job = generics.get_object_or_404(Job, pk=pk)
        stats = {}
        for stage_key, _label in MACRO_STAGE_CHOICES:
            stats[stage_key.lower()] = job.candidate_mappings.filter(macro_stage=stage_key).count()
        stats['total'] = job.candidate_mappings.count()
        return Response(stats)


class JobCollaboratorListCreateView(generics.ListCreateAPIView):
    serializer_class = JobCollaboratorSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [rbac_perm('EDIT_JOBS')()]
        return [rbac_perm('VIEW_JOBS')()]

    def get_queryset(self):
        return JobCollaborator.objects.filter(
            job_id=self.kwargs['pk']
        ).select_related('user')

    def perform_create(self, serializer):
        collab = serializer.save(
            job_id=self.kwargs['pk'],
            added_by=self.request.user,
        )
        log_job_history(
            collab.job, 'collaborator_added', self.request.user,
            f'Collaborator added: {collab.user.full_name}',
            new_value={'user_id': str(collab.user.id), 'name': collab.user.full_name},
        )
        from apps.notifications.utils import notify_collaborator_added
        notify_collaborator_added(collab.job, collab.user, self.request.user)


class JobCollaboratorDeleteView(APIView):
    permission_classes = [rbac_perm('EDIT_JOBS')]

    def delete(self, request, pk, user_id):
        try:
            collab = JobCollaborator.objects.select_related('user', 'job').get(
                job_id=pk, user_id=user_id
            )
            job = collab.job
            user_name = collab.user.full_name
            collab.delete()
            log_job_history(
                job, 'collaborator_removed', request.user,
                f'Collaborator removed: {user_name}',
                previous_value={'user_id': str(user_id), 'name': user_name},
            )
            return Response(status=status.HTTP_204_NO_CONTENT)
        except JobCollaborator.DoesNotExist:
            return Response({'error': 'Collaborator not found'}, status=status.HTTP_404_NOT_FOUND)


class JobHistoryListView(generics.ListAPIView):
    serializer_class = JobHistorySerializer
    permission_classes = [rbac_perm('VIEW_JOBS')]

    def get_queryset(self):
        return JobHistory.objects.filter(
            job_id=self.kwargs['pk']
        ).select_related('changed_by').order_by('-created_at')


class JobExcelReportView(APIView):
    permission_classes = [rbac_perm('VIEW_JOBS')]

    def get(self, request, pk):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from apps.interviews.models import Interview, InterviewFeedback

        job = generics.get_object_or_404(
            Job.objects.select_related(
                'department', 'hiring_manager', 'created_by',
                'requisition', 'requisition__created_by'
            ).prefetch_related('collaborators__user'),
            pk=pk
        )

        mappings = (
            CandidateJobMapping.objects
            .filter(job=job)
            .select_related('candidate')
            .prefetch_related('interviews__feedback', 'interviews__interviewer')
            .order_by('macro_stage', '-stage_updated_at')
        )

        wb = openpyxl.Workbook()

        # ── Styles ────────────────────────────────────────────────────────────
        HEADER_FILL   = PatternFill('solid', fgColor='0058BE')
        SECTION_FILL  = PatternFill('solid', fgColor='DAE2FA')
        HEADER_FONT   = Font(bold=True, color='FFFFFF', size=11)
        SECTION_FONT  = Font(bold=True, color='0058BE', size=10)
        CENTER        = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT          = Alignment(horizontal='left',  vertical='center', wrap_text=True)
        thin          = Side(style='thin', color='C2C6D6')
        BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header_row(ws, row, col_count):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill   = HEADER_FILL
                cell.font   = HEADER_FONT
                cell.alignment = CENTER
                cell.border = BORDER

        def style_section_row(ws, row, col_count):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill   = SECTION_FILL
                cell.font   = SECTION_FONT
                cell.alignment = LEFT
                cell.border = BORDER

        def write_row(ws, row, values):
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.alignment = LEFT
                cell.border    = BORDER

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 1 — Job Overview
        # ══════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Job Overview'
        ws1.column_dimensions['A'].width = 28
        ws1.column_dimensions['B'].width = 50

        req = job.requisition

        job_rows = [
            ('Job Code',           job.job_code),
            ('Job Title',          job.title),
            ('Status',             job.get_status_display()),
            ('Department',         job.department.name if job.department else '—'),
            ('Hiring Manager',     job.hiring_manager.full_name if job.hiring_manager else '—'),
            ('Location',           job.location),
            ('Experience (Min)',   f"{job.experience_min} yrs"),
            ('Experience (Max)',   f"{job.experience_max} yrs"),
            ('Skills Required',    ', '.join(job.skills_required) if job.skills_required else '—'),
            ('Job Description',    job.job_description or '—'),
            ('Positions Filled',   job.positions_filled),
            ('View Count',         job.view_count),
            ('Created By',         job.created_by.full_name if job.created_by else '—'),
            ('Created At',         job.created_at.strftime('%d-%b-%Y')),
        ]

        if req:
            job_rows += [
                ('', ''),
                ('── Requisition ──', ''),
                ('Requisition Code',  req.requisition_code if hasattr(req, 'requisition_code') else '—'),
                ('Purpose',          req.get_purpose_display() if hasattr(req, 'purpose') else '—'),
                ('Purpose Code',     req.purpose_code if hasattr(req, 'purpose_code') else '—'),
                ('Client Name',      req.client_name if hasattr(req, 'client_name') else '—'),
                ('Work Mode',        req.get_work_mode_display() if hasattr(req, 'work_mode') else '—'),
                ('Salary Min (LPA)', str(req.salary_min) if hasattr(req, 'salary_min') and req.salary_min else '—'),
                ('Salary Max (LPA)', str(req.salary_max) if hasattr(req, 'salary_max') and req.salary_max else '—'),
                ('Positions',        req.no_of_positions if hasattr(req, 'no_of_positions') else '—'),
                ('Req. Status',      req.status if hasattr(req, 'status') else '—'),
                ('Req. Created By',  req.created_by.full_name if req.created_by else '—'),
            ]

        collaborators = [c.user.full_name for c in job.collaborators.all()]
        job_rows.append(('Collaborators', ', '.join(collaborators) if collaborators else '—'))

        # Header
        ws1.merge_cells('A1:B1')
        ws1['A1'] = 'JOB OVERVIEW'
        ws1['A1'].fill      = HEADER_FILL
        ws1['A1'].font      = HEADER_FONT
        ws1['A1'].alignment = CENTER
        ws1.row_dimensions[1].height = 24

        for r, (label, value) in enumerate(job_rows, 2):
            ws1.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
            ws1.cell(row=r, column=1).border = BORDER
            ws1.cell(row=r, column=1).fill   = SECTION_FILL
            ws1.cell(row=r, column=2, value=str(value) if value is not None else '—').alignment = LEFT
            ws1.cell(row=r, column=2).border = BORDER
        ws1.row_dimensions[1].height = 24

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 2 — Candidates
        # ══════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Candidates')

        CAND_HEADERS = [
            'S.No', 'Full Name', 'Email', 'Phone', 'Designation',
            'Current Employer', 'Location', 'Experience (yrs)',
            'Current CTC (LPA)', 'Notice Period (days)', 'Expected CTC (LPA)', 'Source',
            'Skills', 'Stage', 'Priority', 'Offer Status',
            'Drop Reason', 'Stage Last Updated',
        ]
        col_widths = [6, 22, 28, 15, 20, 22, 15, 14, 14, 16, 14, 16, 35, 14, 10, 14, 18, 20]
        for i, w in enumerate(col_widths, 1):
            ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

        style_header_row(ws2, 1, len(CAND_HEADERS))
        write_row(ws2, 1, CAND_HEADERS)

        for sno, m in enumerate(mappings, 1):
            c = m.candidate
            write_row(ws2, sno + 1, [
                sno,
                c.full_name,
                c.email,
                c.phone or '—',
                c.designation or '—',
                c.current_employer or '—',
                c.location or '—',
                float(c.total_experience_years) if c.total_experience_years else '—',
                float(c.current_ctc_lakhs) if c.current_ctc_lakhs else '—',
                c.notice_period_days if c.notice_period_days is not None else '—',
                float(c.expected_ctc_lakhs) if c.expected_ctc_lakhs else '—',
                c.get_source_display(),
                ', '.join(c.skills) if c.skills else '—',
                m.macro_stage,
                m.priority,
                m.offer_status or '—',
                m.drop_reason or '—',
                m.stage_updated_at.strftime('%d-%b-%Y %H:%M') if m.stage_updated_at else '—',
            ])

        # ══════════════════════════════════════════════════════════════════════
        # Sheet 3 — Interviews
        # ══════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('Interviews')

        INT_HEADERS = [
            'S.No', 'Candidate Name', 'Round Name', 'Round No.',
            'Interviewer', 'Scheduled At', 'Mode', 'Status',
            'Round Status', 'Round Result', 'Recommendation', 'Overall Score',
        ]
        int_widths = [6, 22, 20, 10, 22, 20, 12, 14, 14, 14, 16, 14]
        for i, w in enumerate(int_widths, 1):
            ws3.column_dimensions[ws3.cell(1, i).column_letter].width = w

        style_header_row(ws3, 1, len(INT_HEADERS))
        write_row(ws3, 1, INT_HEADERS)

        interviews = (
            Interview.objects
            .filter(mapping__job=job)
            .select_related('mapping__candidate', 'interviewer', 'feedback')
            .order_by('mapping__candidate__full_name', 'round_number')
        )

        for sno, iv in enumerate(interviews, 1):
            fb = getattr(iv, 'feedback', None)
            write_row(ws3, sno + 1, [
                sno,
                iv.mapping.candidate.full_name,
                iv.round_name or iv.round_label or '—',
                iv.round_number,
                iv.interviewer.full_name if iv.interviewer else '—',
                iv.scheduled_at.strftime('%d-%b-%Y %H:%M') if iv.scheduled_at else '—',
                iv.mode,
                iv.status,
                iv.round_status or '—',
                iv.round_result or '—',
                fb.recommendation if fb else '—',
                fb.overall_rating if fb else '—',
            ])

        # ── Final response ────────────────────────────────────────────────────
        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"{job.job_code}_{job.title.replace(' ', '_')}_Report.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
